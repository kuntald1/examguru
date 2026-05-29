"""
Vorpet Phase 4 — Export Router
GET /export/results/{exam_id}/pdf    → Rank list PDF
GET /export/results/{exam_id}/excel  → Rank list Excel
GET /export/fees/excel               → Fees report Excel
GET /export/students/excel           → Student list Excel
"""

import io
import datetime
from decimal import Decimal
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

from app.services.db_service import database
from app.dependencies import get_current_institute

router = APIRouter(prefix="/export", tags=["Phase 4 — Export"])


def _val(v):
    if isinstance(v, Decimal): return float(v)
    if isinstance(v, (datetime.datetime, datetime.date)): return str(v)
    return v


def _header_fill():  return PatternFill("solid", fgColor="2E7D32")
def _alt_fill():     return PatternFill("solid", fgColor="F1F8E9")
def _header_font():  return Font(bold=True, color="FFFFFF", size=11)
def _title_font():   return Font(bold=True, size=14, color="1A2E1A")
def _border():
    s = Side(style="thin", color="C8E6C9")
    return Border(left=s, right=s, top=s, bottom=s)


def _style_worksheet(ws, headers: list, col_widths: list):
    """Apply consistent styling to a worksheet"""
    # Title row
    ws.insert_rows(1)
    ws.insert_rows(1)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Vorpet Education"
    title_cell.font  = _title_font()
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(headers))

    # Header row (now row 3)
    header_row = 3
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.fill      = _header_fill()
        c.font      = _header_font()
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border()
    ws.row_dimensions[header_row].height = 22

    # Data rows styling
    for row in ws.iter_rows(min_row=header_row+1):
        for i, cell in enumerate(row):
            cell.border    = _border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.row % 2 == 0:
                cell.fill = _alt_fill()

    # Column widths
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{header_row+1}"


# ── RANK LIST — Excel ─────────────────────────────────────────────────────────

@router.get("/results/{exam_id}/excel")
async def export_results_excel(
    exam_id: int,
    institute: dict = Depends(get_current_institute),
):
    iid = institute["id"]

    exam = await database.fetch_one(
        "SELECT * FROM exams WHERE id=:id AND institute_id=:iid",
        values={"id": exam_id, "iid": iid}
    )
    if not exam:
        raise HTTPException(404, "Exam not found")

    rows = await database.fetch_all("""
        SELECT
            s.name, s.roll_no,
            b.name as batch,
            er.marks_obtained, er.total_marks, er.percentage,
            er.total_correct, er.total_wrong, er.total_skipped,
            er.time_taken_seconds, er.auto_submitted,
            er.submitted_at
        FROM exam_results er
        JOIN exam_access ea ON ea.id = er.exam_access_id
        JOIN students s ON s.id = ea.student_id
        LEFT JOIN batch_students bs ON bs.student_id = s.id
        LEFT JOIN batches b ON bs.batch_id = b.id
        WHERE ea.exam_id = :eid AND s.institute_id = :iid
        ORDER BY er.marks_obtained DESC, er.percentage DESC
    """, values={"eid": exam_id, "iid": iid})

    wb = Workbook()
    ws = wb.active
    ws.title = "Rank List"

    headers   = ["Rank","Name","Roll No","Batch","Marks","Total","Percentage",
                 "Correct","Wrong","Skipped","Time (min)","Auto Submit","Submitted At"]
    col_widths = [6, 24, 12, 16, 8, 8, 12, 8, 8, 8, 11, 12, 20]

    # Set subtitle
    ws.cell(row=2, column=1, value=f"{exam['subject']} — {exam['class_name']} | {exam['school_name']}")

    data_start = 4
    for rank, r in enumerate(rows, 1):
        t = r["time_taken_seconds"] or 0
        ws.append([
            rank,
            r["name"],
            r["roll_no"],
            r["batch"] or "—",
            _val(r["marks_obtained"]),
            _val(r["total_marks"]),
            f"{_val(r['percentage']):.1f}%",
            r["total_correct"],
            r["total_wrong"],
            r["total_skipped"],
            round(t / 60, 1),
            "Yes" if r["auto_submitted"] else "No",
            str(r["submitted_at"])[:16] if r["submitted_at"] else "—",
        ])

    # Colour % column by performance
    pct_col = 7
    for row_idx, r in enumerate(rows, data_start):
        pct = float(r["percentage"] or 0)
        cell = ws.cell(row=row_idx, column=pct_col)
        if pct >= 90:
            cell.fill = PatternFill("solid", fgColor="C8E6C9")
            cell.font = Font(bold=True, color="1B5E20")
        elif pct >= 65:
            cell.fill = PatternFill("solid", fgColor="FFF9C4")
            cell.font = Font(bold=True, color="F57F17")
        else:
            cell.fill = PatternFill("solid", fgColor="FFCDD2")
            cell.font = Font(bold=True, color="B71C1C")

    _style_worksheet(ws, headers, col_widths)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"ranklist_{exam['subject']}_{exam['class_name']}.xlsx".replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )


# ── RANK LIST — PDF ───────────────────────────────────────────────────────────

@router.get("/results/{exam_id}/pdf")
async def export_results_pdf(
    exam_id: int,
    institute: dict = Depends(get_current_institute),
):
    iid = institute["id"]

    exam = await database.fetch_one(
        "SELECT * FROM exams WHERE id=:id AND institute_id=:iid",
        values={"id": exam_id, "iid": iid}
    )
    if not exam:
        raise HTTPException(404, "Exam not found")

    rows = await database.fetch_all("""
        SELECT
            s.name, s.roll_no,
            b.name as batch,
            er.marks_obtained, er.total_marks, er.percentage,
            er.total_correct, er.total_wrong, er.total_skipped,
            er.auto_submitted
        FROM exam_results er
        JOIN exam_access ea ON ea.id = er.exam_access_id
        JOIN students s ON s.id = ea.student_id
        LEFT JOIN batch_students bs ON bs.student_id = s.id
        LEFT JOIN batches b ON bs.batch_id = b.id
        WHERE ea.exam_id = :eid AND s.institute_id = :iid
        ORDER BY er.marks_obtained DESC
    """, values={"eid": exam_id, "iid": iid})

    total = len(rows)
    avg   = round(sum(float(r["percentage"] or 0) for r in rows) / total, 1) if total else 0
    passed = sum(1 for r in rows if float(r["percentage"] or 0) >= 40)

    rows_html = ""
    for rank, r in enumerate(rows, 1):
        pct = float(r["percentage"] or 0)
        if pct >= 90:   bg = "#E8F5E9"; col = "#1B5E20"
        elif pct >= 65: bg = "#FFF9C4"; col = "#F57F17"
        else:           bg = "#FFEBEE"; col = "#B71C1C"
        rows_html += f"""
        <tr style="background:{bg if rank%2==0 else '#fff'}">
          <td style="text-align:center;font-weight:700">{rank}</td>
          <td>{r['name']}</td>
          <td style="text-align:center">{r['roll_no']}</td>
          <td style="text-align:center">{r['batch'] or '—'}</td>
          <td style="text-align:center;font-weight:700">{_val(r['marks_obtained'])}/{_val(r['total_marks'])}</td>
          <td style="text-align:center;font-weight:700;color:{col}">{pct:.1f}%</td>
          <td style="text-align:center">{r['total_correct']}</td>
          <td style="text-align:center">{r['total_wrong']}</td>
        </tr>"""

    html = f"""
    <!DOCTYPE html><html><head><meta charset="UTF-8">
    <style>
      @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap');
      body {{ font-family: Inter, sans-serif; margin: 0; padding: 20px; color: #1A2E1A; font-size: 12px; }}
      .header {{ background: #2E7D32; color: #fff; padding: 18px 24px; border-radius: 8px; margin-bottom: 16px; }}
      .header h1 {{ margin:0; font-size:20px; }} .header p {{ margin:4px 0 0; opacity:0.8; font-size:12px; }}
      .summary {{ display:flex; gap:12px; margin-bottom:16px; }}
      .s-card {{ flex:1; background:#F1F8E9; border:1px solid #C8E6C9; border-radius:8px; padding:12px; text-align:center; }}
      .s-val {{ font-size:22px; font-weight:700; color:#2E7D32; }} .s-lbl {{ font-size:11px; color:#666; margin-top:2px; }}
      table {{ width:100%; border-collapse:collapse; }}
      th {{ background:#2E7D32; color:#fff; padding:8px 10px; font-size:11px; text-align:center; }}
      td {{ padding:7px 10px; border-bottom:1px solid #E8F5E9; }}
      .footer {{ margin-top:16px; text-align:center; font-size:10px; color:#999; }}
    </style></head><body>
    <div class="header">
      <h1>📊 Rank List — {exam['subject']}</h1>
      <p>{exam['school_name']} · {exam['class_name']} · Total Marks: {exam['total_marks']}</p>
    </div>
    <div class="summary">
      <div class="s-card"><div class="s-val">{total}</div><div class="s-lbl">Total Students</div></div>
      <div class="s-card"><div class="s-val">{avg}%</div><div class="s-lbl">Average Score</div></div>
      <div class="s-card"><div class="s-val">{passed}</div><div class="s-lbl">Passed (≥40%)</div></div>
      <div class="s-card"><div class="s-val">{total-passed}</div><div class="s-lbl">Failed</div></div>
    </div>
    <table>
      <thead><tr>
        <th>Rank</th><th>Name</th><th>Roll No</th><th>Batch</th>
        <th>Marks</th><th>Percentage</th><th>Correct</th><th>Wrong</th>
      </tr></thead>
      <tbody>{rows_html}</tbody>
    </table>
    <div class="footer">Generated by Vorpet · {datetime.date.today()}</div>
    </body></html>"""

    from weasyprint import HTML
    pdf_bytes = HTML(string=html).write_pdf()
    fname = f"ranklist_{exam['subject']}_{exam['class_name']}.pdf".replace(" ", "_")
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )


# ── FEES REPORT — Excel ───────────────────────────────────────────────────────

@router.get("/fees/excel")
async def export_fees_excel(
    institute: dict = Depends(get_current_institute),
    month: str = None,
):
    iid = institute["id"]
    if not month:
        month = datetime.date.today().strftime("%Y-%m")

    rows = await database.fetch_all("""
        SELECT
            s.name, s.roll_no, s.phone,
            b.name as batch,
            sfm.month, sfm.amount, sfm.paid,
            (sfm.amount - sfm.paid) as due,
            sfm.status, sfm.note,
            s.admission_date
        FROM student_fees_monthly sfm
        JOIN students s ON sfm.student_id = s.id
        LEFT JOIN batch_students bs ON bs.student_id = s.id
        LEFT JOIN batches b ON bs.batch_id = b.id
        WHERE s.institute_id = :iid AND sfm.month = :month
        ORDER BY b.name, s.name
    """, values={"iid": iid, "month": month})

    wb = Workbook()
    ws = wb.active
    ws.title = f"Fees {month}"
    ws.cell(row=2, column=1, value=f"Fees Report — {month} | {institute['name']}")

    headers = ["Name","Roll No","Phone","Batch","Month","Amount (₹)","Paid (₹)","Due (₹)","Status","Note","Admission Date"]
    col_widths = [22, 10, 14, 16, 10, 12, 12, 12, 12, 20, 16]

    for r in rows:
        ws.append([
            r["name"], r["roll_no"], r["phone"] or "",
            r["batch"] or "—", r["month"],
            float(r["amount"] or 0), float(r["paid"] or 0), float(r["due"] or 0),
            r["status"].upper() if r["status"] else "",
            r["note"] or "",
            str(r["admission_date"]) if r["admission_date"] else "",
        ])

    # Colour status column
    status_col = 9
    data_start  = 4
    for row_idx, r in enumerate(rows, data_start):
        cell = ws.cell(row=row_idx, column=status_col)
        s = (r["status"] or "").lower()
        if s == "paid":
            cell.fill = PatternFill("solid", fgColor="C8E6C9")
            cell.font = Font(bold=True, color="1B5E20")
        elif s == "partial":
            cell.fill = PatternFill("solid", fgColor="FFF9C4")
            cell.font = Font(bold=True, color="F57F17")
        else:
            cell.fill = PatternFill("solid", fgColor="FFCDD2")
            cell.font = Font(bold=True, color="B71C1C")

    _style_worksheet(ws, headers, col_widths)

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    total_amt  = sum(float(r["amount"] or 0) for r in rows)
    total_paid = sum(float(r["paid"] or 0) for r in rows)
    total_due  = sum(float(r["due"] or 0) for r in rows)
    paid_count = sum(1 for r in rows if (r["status"] or "") == "paid")

    ws2.append(["Fees Summary", f"Month: {month}"])
    ws2.append([])
    ws2.append(["Total Billed",  total_amt])
    ws2.append(["Total Collected", total_paid])
    ws2.append(["Total Due",     total_due])
    ws2.append(["Fully Paid Students", paid_count])
    ws2.append(["Total Students", len(rows)])
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="fees_{month}_{institute["name"]}.xlsx"'}
    )


# ── STUDENT LIST — Excel ──────────────────────────────────────────────────────

@router.get("/students/excel")
async def export_students_excel(institute: dict = Depends(get_current_institute)):
    iid = institute["id"]

    rows = await database.fetch_all("""
        SELECT
            s.name, s.roll_no, s.phone, s.email, s.school_name,
            b.name as batch, b.class_name,
            s.admission_date,
            CASE WHEN s.password IS NOT NULL AND s.password != '' THEN 'Set' ELSE 'Not Set' END as password_status
        FROM students s
        LEFT JOIN batch_students bs ON bs.student_id = s.id
        LEFT JOIN batches b ON bs.batch_id = b.id
        WHERE s.institute_id = :iid
        ORDER BY b.name, s.roll_no, s.name
    """, values={"iid": iid})

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.cell(row=2, column=1, value=f"Student List — {institute['name']} | Exported {datetime.date.today()}")

    headers = ["Name","Roll No","Phone","Email","School","Batch","Class","Admission Date","Password"]
    col_widths = [22, 10, 14, 26, 20, 16, 12, 16, 12]

    for r in rows:
        ws.append([
            r["name"], r["roll_no"], r["phone"] or "",
            r["email"] or "", r["school_name"] or "",
            r["batch"] or "—", r["class_name"] or "—",
            str(r["admission_date"]) if r["admission_date"] else "",
            r["password_status"],
        ])

    _style_worksheet(ws, headers, col_widths)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="students_{institute["name"]}.xlsx"'}
    )
